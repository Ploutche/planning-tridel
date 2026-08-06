import json
import re
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

EXCEL_FILE = Path('Classeur1.xlsx')
OUTPUT_FILE = Path('planning.json')
DATE_ROW = 5
FIRST_EMPLOYEE_ROW = 8
TEAM_COLUMN = 1
MATRICULE_COLUMN = 2
NAME_COLUMN = 3
FIRST_DATE_COLUMN = 5


def slugify(value: str) -> str:
    text = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'[^a-zA-Z0-9]+', '-', text).strip('-').lower()
    return text or 'employe'


def normalize_code(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip().upper()


def normalize_date(value, year_hint):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Date Excel : jour 1 = 1899-12-31 avec correction du bug 1900.
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date()
    return None


def main():
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(f'{EXCEL_FILE} est introuvable')

    workbook = load_workbook(EXCEL_FILE, data_only=True, read_only=True)
    worksheet = workbook.active
    year = int(worksheet['A1'].value or datetime.now().year)

    dates = {}
    for column in range(FIRST_DATE_COLUMN, worksheet.max_column + 1):
        parsed = normalize_date(worksheet.cell(DATE_ROW, column).value, year)
        if parsed:
            dates[column] = parsed

    employees = []
    current_team = ''

    for row in range(FIRST_EMPLOYEE_ROW, worksheet.max_row + 1):
        team_value = worksheet.cell(row, TEAM_COLUMN).value
        if team_value:
            current_team = str(team_value).strip()

        matricule = worksheet.cell(row, MATRICULE_COLUMN).value
        name = worksheet.cell(row, NAME_COLUMN).value

        if not name:
            continue

        name = str(name).strip()
        matricule_text = '' if matricule is None else str(matricule).strip()

        if name.upper() == 'VIDE' or matricule_text == '000':
            continue

        schedule = {}
        for column, planning_date in dates.items():
            code = normalize_code(worksheet.cell(row, column).value)
            # Les cases vides restent absentes du JSON pour alléger le fichier.
            if code:
                schedule[planning_date.isoformat()] = code

        employee_id = slugify(f'{name}-{matricule_text}')
        employees.append({
            'id': employee_id,
            'matricule': matricule_text,
            'nom': name,
            'equipe': current_team,
            'planning': schedule,
        })

    output = {
        'annee': year,
        'mise_a_jour': datetime.now().astimezone().isoformat(timespec='seconds'),
        'employes': employees,
    }

    OUTPUT_FILE.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'{OUTPUT_FILE} créé : {len(employees)} employés, {len(dates)} dates.')


if __name__ == '__main__':
    main()
